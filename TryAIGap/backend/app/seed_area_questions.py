"""Seed area questionnaires (module='area') from the Excel kits.

Each kit has a `Cuestionario` (ES) / ` Questionnaire` (EN) sheet with
16 questions in 4 blocks:

    BLOQUE A — <title>      # block header row
    A1 | <question text>    # code in col 1, text in col 2
    ...

Paths are resolved relative to the repo root (this file is
backend/app/seed_area_questions.py, so parents[2] is the repo root).
If a kit file is missing at runtime (e.g. production deploys that do not
ship the Excel files) the seed logs a warning and continues.

NOTE: there is currently NO English marketing kit. The EN translation of
marketing questions falls back to the Spanish text until the kit exists.
"""
import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Question, QuestionTranslation

log = logging.getLogger("app.seed.areas")

REPO_ROOT = Path(__file__).resolve().parents[2]

# area_key -> (spanish kit, english kit or None)
AREA_KITS: Dict[str, Tuple[str, Optional[str]]] = {
    "ventas":      ("Spanish/Kit_Ventas_B2B.xlsx",               "English/EN Sales Kit B2B.xlsx"),
    "marketing":   ("Spanish/Kit_Marketing_DemandGen_B2B.xlsx",  None),  # EN kit does not exist yet
    "servicio":    ("Spanish/Kit_Servicio_Cliente_B2B.xlsx",     "English/EN Customer Service Kit.xlsx"),
    "finanzas":    ("Spanish/Kit_Finanzas_Administracion.xlsx",  "English/EN Finance and Administration.xlsx"),
    "rrhh":        ("Spanish/Kit_RRHH_Talento.xlsx",             "English/EN HR and Talent Kit.xlsx"),
    "operaciones": ("Spanish/Kit_Operaciones_Planta_PYME.xlsx",  "English/EN Plant operations and production.xlsx"),
    "legal":       ("Spanish/Kit_Legal_Compliance.xlsx",         "English/EN Legal and compliance kit.xlsx"),
}

_BLOCK_RE = re.compile(r"^(?:BLOQUE|BLOCK)\s+([A-D])\b", re.IGNORECASE)
_CODE_RE = re.compile(r"^([A-D])(\d)$")


def _cell(value) -> str:
    return str(value).strip() if value is not None else ""


def parse_kit(path: Path) -> Optional[Tuple[Dict[str, str], List[Tuple[str, str]]]]:
    """Parse one kit. Returns ({block_id: block_title}, [(code, text), ...])
    ordered by appearance, or None if the file is missing/unreadable."""
    if not path.is_file():
        log.warning("Area kit not found, skipping: %s", path)
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = next(
                (s for s in wb.sheetnames
                 if s.strip().lower() in ("cuestionario", "questionnaire")),
                None,
            )
            if sheet is None:
                log.warning("No Cuestionario/Questionnaire sheet in %s", path)
                return None
            blocks: Dict[str, str] = {}
            questions: List[Tuple[str, str]] = []
            for row in wb[sheet].iter_rows(values_only=True):
                if not row:
                    continue
                first = _cell(row[0])
                if not first:
                    continue
                m = _BLOCK_RE.match(first)
                if m:
                    blocks[m.group(1).upper()] = first
                    continue
                if _CODE_RE.match(first):
                    text = _cell(row[1]) if len(row) > 1 else ""
                    if text:
                        questions.append((first.upper(), text))
            return blocks, questions
        finally:
            wb.close()
    except Exception:
        log.exception("Failed to parse area kit %s", path)
        return None


async def seed_area_questions(db: AsyncSession) -> None:
    """Idempotently load area questions + ES/EN translations into the DB."""
    existing = (
        await db.execute(
            select(Question.area_key).where(Question.module == "area").distinct()
        )
    ).scalars().all()
    done = set(existing)

    seeded_areas = 0
    for area_key, (es_rel, en_rel) in AREA_KITS.items():
        if area_key in done:
            log.info("Area questions already present for %s. Skipping.", area_key)
            continue

        es = parse_kit(REPO_ROOT / es_rel)
        if es is None:
            continue  # warning already logged; Spanish kit is the source of truth
        es_blocks, es_questions = es

        en_blocks: Dict[str, str] = {}
        en_questions: List[Tuple[str, str]] = []
        if en_rel is not None:
            en = parse_kit(REPO_ROOT / en_rel)
            if en is not None:
                en_blocks, en_questions = en
        else:
            log.warning(
                "No English kit for area '%s' — EN translations fall back to "
                "the Spanish text (known gap: missing EN marketing kit).",
                area_key,
            )
        en_by_code = dict(en_questions)

        for pos, (code, es_text) in enumerate(es_questions, start=1):
            block_id = code[0]
            q = Question(
                module="area",
                area_key=area_key,
                block_id=block_id,
                code=code,
                position=pos,
                dimension=None,
            )
            db.add(q)
            await db.flush()
            db.add(QuestionTranslation(
                question_id=q.id, locale="es",
                text=es_text, block_title=es_blocks.get(block_id),
            ))
            en_text = en_by_code.get(code, es_text)  # ES fallback (marketing gap)
            en_title = en_blocks.get(block_id) or es_blocks.get(block_id)
            db.add(QuestionTranslation(
                question_id=q.id, locale="en",
                text=en_text, block_title=en_title,
            ))
        seeded_areas += 1
        log.info(
            "Seeded area '%s': %d questions in %d blocks (es%s).",
            area_key, len(es_questions), len(es_blocks),
            "+en" if en_questions else ", en=es fallback",
        )

    if seeded_areas:
        print(f"Seeded {seeded_areas} area questionnaire(s).")
    else:
        print("No new area questionnaires to seed.")
